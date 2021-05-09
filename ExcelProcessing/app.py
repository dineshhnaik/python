import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook("transactions.xlsx")
sheet = wb["Sheet1"]

# cell = sheet["A1"]
# print(sheet.max_column)

for row in range(2, sheet.max_row + 1, 1):
    # for col in range(1, sheet.max_column + 1, 1):
    #     print(sheet.cell(row, col).value)
    cell = sheet.cell(row, 3)
    corrected_value = cell.value * 0.9
    corrected_value_cell = sheet.cell(row, 4)
    corrected_value_cell.value = corrected_value

values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'a7' )

wb.save("transactions_updated.xlsx")
